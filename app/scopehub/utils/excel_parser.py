"""
ScopeHub Excel parser - handles reading and writing keyword_strings.xlsx

Expected structure:
- 3 sheets: Pubmed, WOS, Scopus
- Each sheet has sections (A, B, C, etc.)
- Each section has subsections
- Columns: Section, Subsection, Name, Searching Formula, String Combined, Comments
"""
import pandas as pd
import os
from typing import Dict, List, Any


def parse_excel_file(file_path: str) -> tuple[Dict[str, pd.DataFrame], bool, str]:
    """
    Parse keyword_strings.xlsx file with 3 sheets

    Returns:
        Tuple of (sheets_dict, is_valid_format, warning_message)
        - sheets_dict: Dict with keys: 'Pubmed', 'WOS', 'Scopus'
        - is_valid_format: True if file has proper structure with markers
        - warning_message: Warning message if format is old/incorrect
    """
    try:
        excel_file = pd.ExcelFile(file_path)
        sheets = {}
        has_proper_structure = True
        warning_message = ""

        # Expected columns
        expected_columns = [
            'Section', 'Subsection', 'Name', 'Name (Corrected)',
            'Searching Formula', 'String Combined', 'Comments', 'Raw_Output'
        ]

        missing_columns = []
        has_structure_markers = False

        for sheet_name in ['Pubmed', 'WOS', 'Scopus']:
            if sheet_name in excel_file.sheet_names:
                df = pd.read_excel(file_path, sheet_name=sheet_name)

                # Check for missing columns
                for col in expected_columns:
                    if col not in df.columns:
                        if col not in missing_columns:
                            missing_columns.append(col)
                        df[col] = ''

                # Check if file has proper structure markers
                if 'Raw_Output' in df.columns:
                    raw_values = df['Raw_Output'].astype(str).str.strip()
                    if any(marker in raw_values.values for marker in
                           ['__SECTION_HEADER__', '__SUBSECTION_HEADER__', '__COLUMN_HEADER__']):
                        has_structure_markers = True

                # Ensure columns are in expected order
                df = df[expected_columns]

                sheets[sheet_name] = df
            else:
                # Create empty dataframe if sheet doesn't exist
                sheets[sheet_name] = pd.DataFrame(columns=expected_columns)

        # Determine if format is valid
        if missing_columns:
            has_proper_structure = False
            warning_message = f"Uploaded file is missing columns: {', '.join(missing_columns)}. "

        if not has_structure_markers:
            has_proper_structure = False
            warning_message += "File does not have proper structure with section/subsection headers. File will be converted to new format on export."

        return sheets, has_proper_structure, warning_message

    except Exception as e:
        print(f"Error parsing Excel file: {e}")
        return create_empty_sheets(), False, f"Error parsing file: {str(e)}"


def create_empty_sheets() -> Dict[str, pd.DataFrame]:
    """Create empty dataframes for all sheets"""
    empty_df = pd.DataFrame(columns=[
        'Section', 'Subsection', 'Name', 'Name (Corrected)',
        'Searching Formula', 'String Combined', 'Comments', 'Raw_Output'
    ])

    return {
        'Pubmed': empty_df.copy(),
        'WOS': empty_df.copy(),
        'Scopus': empty_df.copy()
    }


def save_to_excel(sheets: Dict[str, pd.DataFrame], file_path: str) -> bool:
    """
    Save sheets back to Excel file with explicit text formatting, visual spacing, and hidden Raw_Output column

    Args:
        sheets: Dict with 'Pubmed', 'WOS', 'Scopus' dataframes
        file_path: Path to save file

    Returns:
        True if successful
    """
    try:
        from openpyxl import Workbook
        from openpyxl.cell.cell import Cell
        from openpyxl.styles import Alignment, PatternFill, Font

        # Create workbook
        wb = Workbook()

        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        for sheet_name, df in sheets.items():
            # Create worksheet
            ws = wb.create_sheet(title=sheet_name)

            # Replace NaN and None with empty strings
            df_clean = df.fillna('')

            # Write header row
            for col_idx, col_name in enumerate(df_clean.columns, start=1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.alignment = Alignment(horizontal='left')

            # Track previous row type for spacing logic
            prev_row_type = None

            # Write data rows
            for row_idx, row_data in enumerate(df_clean.itertuples(index=False), start=2):
                # Get Raw_Output value to determine row type
                raw_output_col_idx = list(df_clean.columns).index('Raw_Output') if 'Raw_Output' in df_clean.columns else -1
                raw_output_value = row_data[raw_output_col_idx] if raw_output_col_idx >= 0 else ''

                for col_idx, (col_name, value) in enumerate(zip(df_clean.columns, row_data), start=1):
                    cell = ws.cell(row=row_idx, column=col_idx)

                    # Convert value to string and handle special cases
                    if pd.isna(value):
                        cell.value = ''
                    else:
                        str_value = str(value)

                        # For formula columns, ALWAYS store as text with explicit data type
                        # This prevents Excel from interpreting search formulas as Excel formulas
                        if col_name in ['Searching Formula', 'String Combined']:
                            # Set as string type explicitly
                            cell.value = str_value
                            cell.data_type = 's'  # Force string type
                            cell.alignment = Alignment(horizontal='left', wrap_text=False)
                        else:
                            cell.value = str_value
                            cell.data_type = 's'  # Force all cells as string type

                # Add visual spacing and formatting based on row type
                if raw_output_value == '__SECTION_HEADER__':
                    # Section headers: light blue background
                    if prev_row_type is not None:  # Not the first section
                        ws.row_dimensions[row_idx].height = 30  # Larger row height for section headers
                    # Light blue background for section headers
                    for col_idx in range(1, len(df_clean.columns) + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")

                # Subsection headers: light yellow background
                elif raw_output_value == '__SUBSECTION_HEADER__':
                    ws.row_dimensions[row_idx].height = 25  # Medium row height for subsection headers
                    # Light yellow background for subsection headers
                    for col_idx in range(1, len(df_clean.columns) + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")

                # Column headers within subsections: bold font, light gray background
                elif raw_output_value == '__COLUMN_HEADER__':
                    ws.row_dimensions[row_idx].height = 20  # Standard row height for column headers
                    # Light gray background and bold font for column headers
                    for col_idx in range(1, len(df_clean.columns) + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center')

                # Combined string rows: spacing for subsection separator
                elif raw_output_value == '__COMBINED_STRING_ROW__':
                    ws.row_dimensions[row_idx].height = 35  # Spacing between subsections

                # Separator rows: EXTRA LARGE spacing between sections
                elif raw_output_value == '__SEPARATOR_ROW__':
                    ws.row_dimensions[row_idx].height = 50  # Extra large spacing between sections

                prev_row_type = raw_output_value

            # Auto-adjust column widths and hide Raw_Output
            for col_idx, col_name in enumerate(df_clean.columns, start=1):
                col_letter = ws.cell(row=1, column=col_idx).column_letter

                if col_name == 'Raw_Output':
                    # Hide the Raw_Output column (width = 0)
                    ws.column_dimensions[col_letter].width = 0
                    ws.column_dimensions[col_letter].hidden = True
                elif col_name in ['Searching Formula', 'String Combined']:
                    ws.column_dimensions[col_letter].width = 50
                elif col_name == 'Comments':
                    ws.column_dimensions[col_letter].width = 30
                elif col_name == 'Name (Corrected)':
                    ws.column_dimensions[col_letter].width = 25
                elif col_name in ['Section', 'Subsection']:
                    ws.column_dimensions[col_letter].width = 12
                else:
                    ws.column_dimensions[col_letter].width = 15

        # Save workbook
        wb.save(file_path)
        return True

    except Exception as e:
        print(f"Error saving Excel file: {e}")
        return False


def get_sections(df: pd.DataFrame) -> List[str]:
    """Get unique sections from dataframe"""
    if 'Section' in df.columns:
        return sorted(df['Section'].dropna().unique().tolist())
    return []


def get_subsections(df: pd.DataFrame, section: str) -> List[str]:
    """Get subsections for a specific section"""
    if 'Section' in df.columns and 'Subsection' in df.columns:
        section_df = df[df['Section'] == section]
        return sorted(section_df['Subsection'].dropna().unique().tolist())
    return []


def add_entry(df: pd.DataFrame, section: str, subsection: str, name: str,
              searching_formula: str = '', string_combined: str = '',
              comments: str = '') -> pd.DataFrame:
    """Add new entry to dataframe"""
    new_row = {
        'Section': section,
        'Subsection': subsection,
        'Name': name,
        'Searching Formula': searching_formula,
        'String Combined': string_combined,
        'Comments': comments
    }

    return pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)


def update_entry(df: pd.DataFrame, index: int, column: str, value: str) -> pd.DataFrame:
    """Update specific entry in dataframe"""
    if index < len(df) and column in df.columns:
        df.at[index, column] = value
    return df


def get_entries_by_section(df: pd.DataFrame, section: str, subsection: str = None) -> pd.DataFrame:
    """Get all entries for a section (and optionally subsection)"""
    filtered = df[df['Section'] == section]

    if subsection:
        filtered = filtered[filtered['Subsection'] == subsection]

    return filtered


def sheets_to_unified(sheets: Dict[str, pd.DataFrame]) -> List[Dict[str, Any]]:
    """
    Convert 3 separate sheets into unified structure.

    Unified structure has one entry per Name with formulas for all 3 databases:
    {
        'Section': 'A',
        'Subsection': 'A1',
        'Name': 'topic',
        'Pubmed_Formula': '...',
        'WOS_Formula': '...',
        'Scopus_Formula': '...',
        'Comments': '...',
        'Pubmed_Raw': '...',  # Raw output if there was an error
        'WOS_Raw': '...',
        'Scopus_Raw': '...'
    }

    We assume that Section, Subsection, Name are the same across all 3 sheets.
    We'll use Pubmed as the base and merge formulas from WOS and Scopus.

    Skips:
    - String Combined rows (empty subsection)
    - Empty separator rows (empty section)
    """
    pubmed_df = sheets.get('Pubmed', pd.DataFrame())
    wos_df = sheets.get('WOS', pd.DataFrame())
    scopus_df = sheets.get('Scopus', pd.DataFrame())

    # Start with Pubmed as base
    if pubmed_df.empty:
        print("DEBUG: Pubmed dataframe is empty")
        return []

    print(f"DEBUG: Pubmed has {len(pubmed_df)} rows, columns: {pubmed_df.columns.tolist()}")

    unified = []

    for _, row in pubmed_df.iterrows():
        section = str(row.get('Section', '')).strip()
        subsection = str(row.get('Subsection', '')).strip()
        name = str(row.get('Name', '')).strip() if pd.notna(row.get('Name', '')) else ''

        # Skip special rows (marked in Raw_Output column)
        # We use markers to distinguish:
        # - Section headers: Raw_Output = '__SECTION_HEADER__'
        # - Subsection headers: Raw_Output = '__SUBSECTION_HEADER__'
        # - Column headers: Raw_Output = '__COLUMN_HEADER__'
        # - Combined String rows: Raw_Output = '__COMBINED_STRING_ROW__'
        # - Separator rows: Raw_Output = '__SEPARATOR_ROW__'
        # These are NOT data entries and should be skipped when loading
        # We use Raw_Output instead of Comments so markers are less visible to users
        raw_output = str(row.get('Raw_Output', '')).strip()
        if raw_output in ['__SECTION_HEADER__', '__SUBSECTION_HEADER__', '__COLUMN_HEADER__', '__COMBINED_STRING_ROW__', '__SEPARATOR_ROW__']:
            continue

        # Skip rows where Section starts with '===' (old-style headers)
        if section.startswith('==='):
            continue

        # IMPORTANT: Don't skip entries with empty name!
        # These are new entries waiting for user to fill in.
        # They must be preserved so they appear in the UI.

        # Find matching entries in WOS and Scopus
        # Handle empty subsections (when section has only one subsection)
        # Filter out header rows before matching
        wos_clean = wos_df[~wos_df['Section'].astype(str).str.startswith('===')]
        scopus_clean = scopus_df[~scopus_df['Section'].astype(str).str.startswith('===')]

        wos_match = wos_clean[
            (wos_clean['Section'] == section) &
            (wos_clean['Subsection'].fillna('').str.strip() == subsection) &
            (wos_clean['Name'] == name)
        ]

        scopus_match = scopus_clean[
            (scopus_clean['Section'] == section) &
            (scopus_clean['Subsection'].fillna('').str.strip() == subsection) &
            (scopus_clean['Name'] == name)
        ]

        # Helper function to safely get and convert values
        def safe_get(value):
            """Convert pandas/numpy values to Python native types, handle NaN"""
            if value is None or pd.isna(value):
                return ''
            if isinstance(value, (int, float, bool)):
                if isinstance(value, float) and str(value) == 'nan':
                    return ''
                return value
            return str(value)

        entry = {
            'Section': section,
            'Subsection': subsection,
            'Name': name,
            'Name_Corrected': safe_get(row.get('Name (Corrected)', '')),
            'Pubmed_Formula': safe_get(row.get('Searching Formula', '')),
            'WOS_Formula': safe_get(wos_match.iloc[0]['Searching Formula']) if len(wos_match) > 0 else '',
            'Scopus_Formula': safe_get(scopus_match.iloc[0]['Searching Formula']) if len(scopus_match) > 0 else '',
            'Comments': safe_get(row.get('Comments', '')),
            'Pubmed_Raw': safe_get(row.get('Raw_Output', '')),
            'WOS_Raw': safe_get(wos_match.iloc[0]['Raw_Output']) if len(wos_match) > 0 else '',
            'Scopus_Raw': safe_get(scopus_match.iloc[0]['Raw_Output']) if len(scopus_match) > 0 else ''
        }

        unified.append(entry)

    print(f"DEBUG: Converted {len(unified)} entries to unified format")
    return unified


def unified_to_sheets(unified: List[Dict[str, Any]]) -> Dict[str, pd.DataFrame]:
    """
    Convert unified structure back to 3 separate sheets.

    New structure:
    - Section header row (marks start of section)
    - Subsection entries with multiple Names
    - Combined String after EACH subsection
    - If section has only one subsection, don't use number (B instead of B1)
    - Empty row separator between sections

    Example:
    === SECTION A ===
    A | A1 | Name1 | formula1 | | ...
    A | A1 | Name2 | formula2 | | ...
    A | A1 |       |          | (f1) OR (f2) |
    A | A2 | Name3 | formula3 | | ...
    A | A2 |       |          | (f3) |
      |    |       |          | |  (separator)

    === SECTION B === (only one subsection - no B1)
    B |    | Name4 | formula4 | | ...
    B |    | Name5 | formula5 | | ...
    B |    |       |          | (f4) OR (f5) |
      |    |       |          | |  (separator)
    """
    # Group entries by section and subsection
    sections = {}
    for entry in unified:
        section = entry.get('Section', '')
        subsection = entry.get('Subsection', '')

        if section not in sections:
            sections[section] = {}

        if subsection not in sections[section]:
            sections[section][subsection] = []

        sections[section][subsection].append(entry)

    pubmed_data = []
    wos_data = []
    scopus_data = []

    # Process each section
    for section_name in sorted(sections.keys()):
        subsections = sections[section_name]

        # Add section header row
        pubmed_data.append({
            'Section': f'=== SECTION {section_name} ===',
            'Subsection': '',
            'Name': '',
            'Name (Corrected)': '',
            'Searching Formula': '',
            'String Combined': '',
            'Comments': '',
            'Raw_Output': '__SECTION_HEADER__'  # Marker in Raw_Output
        })

        wos_data.append({
            'Section': f'=== SECTION {section_name} ===',
            'Subsection': '',
            'Name': '',
            'Name (Corrected)': '',
            'Searching Formula': '',
            'String Combined': '',
            'Comments': '',
            'Raw_Output': '__SECTION_HEADER__'  # Marker in Raw_Output
        })

        scopus_data.append({
            'Section': f'=== SECTION {section_name} ===',
            'Subsection': '',
            'Name': '',
            'Name (Corrected)': '',
            'Searching Formula': '',
            'String Combined': '',
            'Comments': '',
            'Raw_Output': '__SECTION_HEADER__'  # Marker in Raw_Output
        })

        # ALWAYS use subsection numbers in Excel for consistency
        # Even if there's only one subsection, we keep the number (A1, B1, etc.)
        # This prevents confusion and allows proper tracking when adding more subsections

        # Process each subsection
        for subsection_name in sorted(subsections.keys()):
            subsection_entries = subsections[subsection_name]

            # Always use the full subsection name (e.g., A1, A2, B1)
            displayed_subsection = subsection_name

            # Add subsection header row
            pubmed_data.append({
                'Section': '',
                'Subsection': f'=== SUBSECTION {subsection_name} ===',
                'Name': '',
                'Name (Corrected)': '',
                'Searching Formula': '',
                'String Combined': '',
                'Comments': '',
                'Raw_Output': '__SUBSECTION_HEADER__'  # Marker for subsection header
            })

            wos_data.append({
                'Section': '',
                'Subsection': f'=== SUBSECTION {subsection_name} ===',
                'Name': '',
                'Name (Corrected)': '',
                'Searching Formula': '',
                'String Combined': '',
                'Comments': '',
                'Raw_Output': '__SUBSECTION_HEADER__'  # Marker for subsection header
            })

            scopus_data.append({
                'Section': '',
                'Subsection': f'=== SUBSECTION {subsection_name} ===',
                'Name': '',
                'Name (Corrected)': '',
                'Searching Formula': '',
                'String Combined': '',
                'Comments': '',
                'Raw_Output': '__SUBSECTION_HEADER__'  # Marker for subsection header
            })

            # Add column header row after subsection header
            pubmed_data.append({
                'Section': 'Section',
                'Subsection': 'Subsection',
                'Name': 'Name',
                'Name (Corrected)': 'Name (Corrected)',
                'Searching Formula': 'Searching Formula',
                'String Combined': 'String Combined',
                'Comments': 'Comments',
                'Raw_Output': '__COLUMN_HEADER__'  # Marker for column header row
            })

            wos_data.append({
                'Section': 'Section',
                'Subsection': 'Subsection',
                'Name': 'Name',
                'Name (Corrected)': 'Name (Corrected)',
                'Searching Formula': 'Searching Formula',
                'String Combined': 'String Combined',
                'Comments': 'Comments',
                'Raw_Output': '__COLUMN_HEADER__'  # Marker for column header row
            })

            scopus_data.append({
                'Section': 'Section',
                'Subsection': 'Subsection',
                'Name': 'Name',
                'Name (Corrected)': 'Name (Corrected)',
                'Searching Formula': 'Searching Formula',
                'String Combined': 'String Combined',
                'Comments': 'Comments',
                'Raw_Output': '__COLUMN_HEADER__'  # Marker for column header row
            })

            # Collect formulas for this subsection's Combined String
            pubmed_formulas = []
            wos_formulas = []
            scopus_formulas = []

            # Add all entries for this subsection
            for entry in subsection_entries:
                name = entry.get('Name', '')
                name_corrected = entry.get('Name_Corrected', '')  # LLM-corrected name
                comments = entry.get('Comments', '')

                pubmed_formula = entry.get('Pubmed_Formula', '')
                wos_formula = entry.get('WOS_Formula', '')
                scopus_formula = entry.get('Scopus_Formula', '')

                pubmed_raw = entry.get('Pubmed_Raw', '')
                wos_raw = entry.get('WOS_Raw', '')
                scopus_raw = entry.get('Scopus_Raw', '')

                # Collect non-empty formulas for combined string
                if pubmed_formula and pubmed_formula.strip():
                    pubmed_formulas.append(pubmed_formula)
                if wos_formula and wos_formula.strip():
                    wos_formulas.append(wos_formula)
                if scopus_formula and scopus_formula.strip():
                    scopus_formulas.append(scopus_formula)

                # Add entry rows
                pubmed_data.append({
                    'Section': section_name,
                    'Subsection': displayed_subsection,
                    'Name': name,
                    'Name (Corrected)': name_corrected,  # LLM-corrected name if available
                    'Searching Formula': pubmed_formula,
                    'String Combined': '',
                    'Comments': comments,
                    'Raw_Output': pubmed_raw
                })

                wos_data.append({
                    'Section': section_name,
                    'Subsection': displayed_subsection,
                    'Name': name,
                    'Name (Corrected)': name_corrected,  # LLM-corrected name if available
                    'Searching Formula': wos_formula,
                    'String Combined': '',
                    'Comments': comments,
                    'Raw_Output': wos_raw
                })

                scopus_data.append({
                    'Section': section_name,
                    'Subsection': displayed_subsection,
                    'Name': name,
                    'Name (Corrected)': name_corrected,  # LLM-corrected name if available
                    'Searching Formula': scopus_formula,
                    'String Combined': '',
                    'Comments': comments,
                    'Raw_Output': scopus_raw
                })

            # Add Combined String row for this subsection
            # Smart combining based on database syntax

            # PubMed: Simple OR joining (no special prefix)
            pubmed_combined = ' OR '.join([f"({f})" for f in pubmed_formulas]) if pubmed_formulas else ''

            # WOS: Extract TS= prefix if all formulas use it
            if wos_formulas:
                print(f"DEBUG WOS combining: {len(wos_formulas)} formulas")
                for i, f in enumerate(wos_formulas):
                    print(f"  Formula {i}: '{f[:80]}...'")

                # First strip outer parentheses from all formulas to detect prefix
                stripped_formulas = []
                for f in wos_formulas:
                    f_stripped = f.strip()
                    # Remove outer parentheses if present
                    if f_stripped.startswith('(') and f_stripped.endswith(')'):
                        f_stripped = f_stripped[1:-1].strip()
                    stripped_formulas.append(f_stripped)
                    print(f"  Stripped: '{f_stripped[:80]}...'")

                # Check if all formulas start with common prefix like TS=, TI=, AU=, etc.
                wos_prefix = None
                for prefix in ['TS=', 'TI=', 'AU=', 'SO=', 'AB=']:
                    if all(f.startswith(prefix) for f in stripped_formulas):
                        wos_prefix = prefix
                        break

                print(f"  Detected WOS prefix: {wos_prefix}")

                if wos_prefix:
                    # Extract content from parentheses and combine
                    contents = []
                    for formula in stripped_formulas:
                        # Remove prefix and extract content
                        without_prefix = formula[len(wos_prefix):]
                        # Remove outer parentheses if present
                        if without_prefix.startswith('(') and without_prefix.endswith(')'):
                            without_prefix = without_prefix[1:-1]
                        contents.append(without_prefix)

                    # Combine with single prefix
                    wos_combined = f"{wos_prefix}({' OR '.join([f'({c})' for c in contents])})"
                else:
                    # No common prefix - simple OR joining
                    wos_combined = ' OR '.join([f"({f})" for f in wos_formulas])
            else:
                wos_combined = ''

            # Scopus: Extract TITLE-ABS-KEY prefix if all formulas use it
            if scopus_formulas:
                print(f"DEBUG Scopus combining: {len(scopus_formulas)} formulas")
                for i, f in enumerate(scopus_formulas):
                    print(f"  Formula {i}: '{f[:80]}...'")

                # First strip outer parentheses from all formulas to detect prefix
                stripped_formulas = []
                for f in scopus_formulas:
                    f_stripped = f.strip()
                    # Remove outer parentheses if present
                    if f_stripped.startswith('(') and f_stripped.endswith(')'):
                        f_stripped = f_stripped[1:-1].strip()
                    stripped_formulas.append(f_stripped)
                    print(f"  Stripped: '{f_stripped[:80]}...'")

                # Check if all formulas start with TITLE-ABS-KEY
                has_prefix = all(f.startswith('TITLE-ABS-KEY') for f in stripped_formulas)
                print(f"  All have TITLE-ABS-KEY prefix: {has_prefix}")

                if has_prefix:
                    # Extract content from parentheses and combine
                    contents = []
                    for formula in stripped_formulas:
                        # Remove "TITLE-ABS-KEY" and extract content
                        without_prefix = formula[13:]  # len('TITLE-ABS-KEY') = 13
                        # Remove outer parentheses if present
                        if without_prefix.startswith('(') and without_prefix.endswith(')'):
                            without_prefix = without_prefix[1:-1]
                        contents.append(without_prefix)

                    # Combine with single prefix
                    scopus_combined = f"TITLE-ABS-KEY({' OR '.join([f'({c})' for c in contents])})"
                else:
                    # No common prefix - simple OR joining
                    scopus_combined = ' OR '.join([f"({f})" for f in scopus_formulas])
            else:
                scopus_combined = ''

            pubmed_data.append({
                'Section': section_name,
                'Subsection': displayed_subsection,
                'Name': '',
                'Name (Corrected)': '',
                'Searching Formula': '',
                'String Combined': pubmed_combined,
                'Comments': '',
                'Raw_Output': '__COMBINED_STRING_ROW__'  # Marker to distinguish from empty entry rows
            })

            wos_data.append({
                'Section': section_name,
                'Subsection': displayed_subsection,
                'Name': '',
                'Name (Corrected)': '',
                'Searching Formula': '',
                'String Combined': wos_combined,
                'Comments': '',
                'Raw_Output': '__COMBINED_STRING_ROW__'  # Marker to distinguish from empty entry rows
            })

            scopus_data.append({
                'Section': section_name,
                'Subsection': displayed_subsection,
                'Name': '',
                'Name (Corrected)': '',
                'Searching Formula': '',
                'String Combined': scopus_combined,
                'Comments': '',
                'Raw_Output': '__COMBINED_STRING_ROW__'  # Marker to distinguish from empty entry rows
            })

        # Add separator row between sections
        pubmed_data.append({
            'Section': '',
            'Subsection': '',
            'Name': '',
            'Name (Corrected)': '',
            'Searching Formula': '',
            'String Combined': '',
            'Comments': '',
            'Raw_Output': '__SEPARATOR_ROW__'  # Marker
        })

        wos_data.append({
            'Section': '',
            'Subsection': '',
            'Name': '',
            'Name (Corrected)': '',
            'Searching Formula': '',
            'String Combined': '',
            'Comments': '',
            'Raw_Output': '__SEPARATOR_ROW__'  # Marker
        })

        scopus_data.append({
            'Section': '',
            'Subsection': '',
            'Name': '',
            'Name (Corrected)': '',
            'Searching Formula': '',
            'String Combined': '',
            'Comments': '',
            'Raw_Output': '__SEPARATOR_ROW__'  # Marker
        })

    return {
        'Pubmed': pd.DataFrame(pubmed_data),
        'WOS': pd.DataFrame(wos_data),
        'Scopus': pd.DataFrame(scopus_data)
    }
